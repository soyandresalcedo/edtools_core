# Copyright (c) 2026, EdTools and contributors
# Importación masiva de Course Enrollment desde Excel/CSV (ID, SEMESTER, COURSE).

from __future__ import annotations

import csv
import html
import re
from collections import defaultdict
from typing import Any, Callable

import datetime

import frappe
from frappe import _
from frappe.utils import getdate, nowdate

from edtools_core.grade_import import (
	SEMESTER_SUFFIX_TO_TERM,
	_find_column_index,
	_resolve_course,
	_resolve_file_path,
	get_student_name_by_id,
	semester_to_academic_year_and_term,
)

REQUIRED_COLUMNS = ["ID", "SEMESTER", "COURSE"]
OPTIONAL_COLUMNS = ["ENROLLMENT DATE"]

SEMESTER_RE = re.compile(r"^\d{6}$")


def _plain_user_message_from_exception(exc: BaseException, max_len: int = 220) -> str:
	"""Quita etiquetas HTML y compacta espacios para tablas de importación (Text Editor)."""
	s = str(exc) if exc else ""
	s = re.sub(r"<[^>]+>", " ", s)
	s = html.unescape(s)
	s = " ".join(s.split()).strip()
	if len(s) > max_len:
		return s[: max_len - 1] + "…"
	return s


def coerce_enrollment_date_str(val) -> str | None:
	"""
	Normaliza fechas desde el formulario Single (Date puede ser date/datetime/str),
	desde Excel (openpyxl devuelve datetime.date) o desde CSV (str).
	Devuelve 'YYYY-MM-DD' o None.
	"""
	if val is None or val == "":
		return None
	if isinstance(val, str):
		s = val.strip()
		if not s:
			return None
		try:
			return str(getdate(s))
		except Exception:
			return None
	if isinstance(val, datetime.datetime):
		return str(val.date())
	if isinstance(val, datetime.date):
		return str(val)
	try:
		return str(getdate(val))
	except Exception:
		return None


def _parse_import_csv(file_path: str) -> tuple[dict[str, int], list[dict[str, Any]]]:
	with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
		reader = csv.reader(f)
		rows_list = list(reader)
	if not rows_list:
		return {}, []
	header_row = [str(c).strip() for c in rows_list[0]]
	col_index: dict[str, int] = {}
	for col_name in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
		idx = _find_column_index(header_row, col_name)
		if idx is not None:
			col_index[col_name] = idx
	data_rows: list[dict[str, Any]] = []
	for r in rows_list[1:]:
		row_dict: dict[str, Any] = {}
		for name, idx in col_index.items():
			if idx < len(r):
				row_dict[name] = (r[idx] or "").strip()
			else:
				row_dict[name] = ""
		data_rows.append(row_dict)
	return col_index, data_rows


def _parse_import_xlsx(file_path: str) -> tuple[dict[str, int], list[dict[str, Any]]]:
	from openpyxl import load_workbook

	wb = load_workbook(file_path, read_only=True, data_only=True)
	ws = wb.active
	rows_list: list[list[str]] = []
	for row in ws.iter_rows(values_only=True):
		rows_list.append([str(c) if c is not None else "" for c in row])
	wb.close()
	if not rows_list:
		return {}, []
	header_row = [str(c).strip() for c in rows_list[0]]
	col_index: dict[str, int] = {}
	for col_name in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
		idx = _find_column_index(header_row, col_name)
		if idx is not None:
			col_index[col_name] = idx
	data_rows: list[dict[str, Any]] = []
	for r in rows_list[1:]:
		row_dict: dict[str, Any] = {}
		for name, idx in col_index.items():
			if idx < len(r):
				val = r[idx]
				if name == "ENROLLMENT DATE" and val is not None and val != "":
					coerced = coerce_enrollment_date_str(val)
					row_dict[name] = coerced or ""
				else:
					row_dict[name] = (str(val).strip() if val is not None else "")
			else:
				row_dict[name] = ""
		data_rows.append(row_dict)
	return col_index, data_rows


def parse_import_file(file_path: str) -> tuple[dict[str, int], list[dict[str, Any]]]:
	path_lower = (file_path or "").lower()
	if path_lower.endswith(".csv"):
		return _parse_import_csv(file_path)
	if path_lower.endswith(".xlsx") or path_lower.endswith(".xls"):
		return _parse_import_xlsx(file_path)
	return {}, []


def validate_import_format(file_path: str) -> tuple[bool, list[dict[str, Any]]]:
	"""Validación previa: columnas, filas, formato SEMESTER. No toca PE ni Moodle."""
	errors: list[dict[str, Any]] = []
	resolved = _resolve_file_path(file_path)
	if not resolved:
		errors.append({"row": None, "message": _("Se requiere un archivo Excel (.xlsx) o CSV.")})
		return False, errors
	file_path = resolved
	if not (
		file_path.lower().endswith(".csv")
		or file_path.lower().endswith(".xlsx")
		or file_path.lower().endswith(".xls")
	):
		errors.append({"row": None, "message": _("El archivo debe ser Excel (.xlsx) o CSV.")})
		return False, errors

	col_index, data_rows = parse_import_file(file_path)
	missing = [c for c in REQUIRED_COLUMNS if c not in col_index]
	if missing:
		errors.append(
			{
				"row": None,
				"message": _("Faltan columnas requeridas: {0}").format(", ".join(missing)),
			}
		)
		return False, errors

	if not data_rows:
		errors.append({"row": None, "message": _("El archivo no tiene filas de datos.")})
		return False, errors

	for i, row in enumerate(data_rows):
		row_num = i + 2
		student_id = (row.get("ID") or "").strip()
		if not student_id:
			errors.append(
				{
					"row": row_num,
					"message": _("Fila {0}: ID de estudiante no puede estar vacío.").format(row_num),
				}
			)
		semester = (row.get("SEMESTER") or "").strip().replace(" ", "")
		if not semester:
			errors.append(
				{
					"row": row_num,
					"message": _("Fila {0}: SEMESTER no puede estar vacío.").format(row_num),
				}
			)
		elif not SEMESTER_RE.match(semester):
			errors.append(
				{
					"row": row_num,
					"message": _("Fila {0}: SEMESTER debe ser 6 dígitos numéricos (ej. 202601 o 202600).").format(
						row_num
					),
				}
			)
		else:
			suffix = semester[-2:]
			if suffix not in SEMESTER_SUFFIX_TO_TERM:
				errors.append(
					{
						"row": row_num,
						"message": _(
							"Fila {0}: SEMESTER inválido; use 00 (Homologado) o 01–06 (semestres), ej. 202600 o 202601."
						).format(row_num),
					}
				)
			else:
				parsed = semester_to_academic_year_and_term(semester)
				if parsed:
					_year_part, term_name = parsed
					if not frappe.db.exists("Academic Term", term_name):
						errors.append(
							{
								"row": row_num,
								"message": _(
									"Fila {0}: no existe Academic Term '{1}' en el sistema."
								).format(row_num, term_name),
							}
						)
		course = (row.get("COURSE") or "").strip()
		if not course:
			errors.append(
				{
					"row": row_num,
					"message": _("Fila {0}: COURSE no puede estar vacío.").format(row_num),
				}
			)

	if errors:
		return False, errors
	return True, []


def build_import_student_group_name(course_doc, year: str, term_label: str) -> str:
	course_name = (course_doc.course_name or "").strip()
	if " - " in course_name:
		short = course_name.split(" - ", 1)[0].strip()
		title = course_name.split(" - ", 1)[1].strip()
	else:
		short = (getattr(course_doc, "short_name", None) or course_name or course_doc.name).strip()
		title = course_name or short
	return f"Course - {short} - {title} - {year} ({term_label})"


def _unique_preserve_order(items: list[str]) -> list[str]:
	seen: set[str] = set()
	out: list[str] = []
	for x in items:
		if x not in seen:
			seen.add(x)
			out.append(x)
	return out


def ensure_student_group_for_course_import(
	course_frappe: str,
	academic_year: str,
	academic_term_name: str,
	term_label: str,
	year: str,
	student_names: list[str],
) -> str:
	course_doc = frappe.get_cached_doc("Course", course_frappe)
	group_name = build_import_student_group_name(course_doc, year, term_label)
	student_names = _unique_preserve_order([s for s in student_names if s])

	if frappe.db.exists("Student Group", group_name):
		doc = frappe.get_doc("Student Group", group_name)
		existing = {r.student for r in (doc.students or [])}
		max_roll = 0
		for r in doc.students or []:
			try:
				max_roll = max(max_roll, int(r.group_roll_number or 0))
			except (TypeError, ValueError):
				pass
		appended = False
		for stu in student_names:
			if stu not in existing:
				max_roll += 1
				student_name_title = frappe.db.get_value("Student", stu, "student_name") or stu
				doc.append(
					"students",
					{
						"student": stu,
						"student_name": student_name_title,
						"group_roll_number": max_roll,
						"active": 1,
					},
				)
				existing.add(stu)
				appended = True
		if appended:
			# validate_course en Student Group exige que el alumno ya esté en el curso en el PE;
			# en import masivo aún no existe Course Enrollment → omitir validación.
			doc.flags.ignore_validate = True
			doc.save(ignore_permissions=True)
		return group_name

	doc = frappe.new_doc("Student Group")
	doc.academic_year = academic_year
	doc.group_based_on = "Course"
	doc.student_group_name = group_name
	doc.academic_term = academic_term_name
	doc.course = course_frappe
	doc.program = None
	doc.max_strength = 0
	for i, stu in enumerate(student_names, 1):
		student_name_title = frappe.db.get_value("Student", stu, "student_name") or stu
		doc.append(
			"students",
			{
				"student": stu,
				"student_name": student_name_title,
				"group_roll_number": i,
				"active": 1,
			},
		)
	doc.flags.ignore_validate = True
	doc.insert(ignore_permissions=True)
	return doc.name


def get_unique_program_enrollment(student: str, academic_year: str) -> tuple[str | None, str | None, str | None]:
	"""
	Retorna (pe_name, pe_year, error).

	Prioriza el PE submitted del año solicitado y, si no existe, usa el PE
	submitted más reciente del estudiante para permitir cursos de años posteriores.
	"""

	def _pick_latest(filters: dict[str, Any]) -> list[dict[str, Any]]:
		return frappe.get_all(
			"Program Enrollment",
			filters=filters,
			fields=["name", "academic_year"],
			order_by="modified desc",
			limit_page_length=1,
		)

	pes = _pick_latest({"student": student, "academic_year": academic_year, "docstatus": 1})
	if not pes:
		pes = _pick_latest({"student": student, "docstatus": 1})

	if not pes:
		return None, None, _("Estudiante sin Program Enrollment activo")

	pe = pes[0]
	return pe.get("name"), pe.get("academic_year"), None


def process_enrollments(
	file_path: str,
	default_enrollment_date: str | None = None,
	progress_callback: Callable | None = None,
) -> dict[str, Any]:
	results: list[dict[str, Any]] = []

	def _add_result(
		*,
		row: int | None,
		student_id: str = "",
		student: str = "",
		course_input: str = "",
		course: str = "",
		academic_term: str = "",
		status: str = "",
		detail: str = "",
	):
		results.append(
			{
				"row": row,
				"student_id": student_id or "",
				"student": student or "",
				"course_input": course_input or "",
				"course": course or "",
				"academic_term": academic_term or "",
				"status": status or "",
				"detail": detail or "",
			}
		)

	out: dict[str, Any] = {
		"success": False,
		"validation_errors": [],
		"summary": {
			"course_enrollments_created": 0,
			"duplicates": 0,
			"rows_processed_ok": 0,
			"rows_with_errors": 0,
			"student_groups_created_or_updated": 0,
		},
		"errors": [],
		"results": results,
	}

	ok, validation_errors = validate_import_format(file_path)
	if not ok:
		out["validation_errors"] = validation_errors
		for e in validation_errors:
			_add_result(
				row=e.get("row"),
				status="ErrorValidacion",
				detail=e.get("message", ""),
			)
		return out

	resolved = _resolve_file_path(file_path)
	if not resolved:
		out["validation_errors"] = [{"row": None, "message": _("No se pudo leer el archivo.")}]
		return out

	_unused_col_index, data_rows = parse_import_file(resolved)
	coerced_default = coerce_enrollment_date_str(default_enrollment_date)
	default_date = coerced_default or nowdate()

	groups: dict[tuple[str, str, str], list[tuple[int, str, str, str]]] = defaultdict(list)
	for i, row in enumerate(data_rows):
		row_num = i + 2
		semester = (row.get("SEMESTER") or "").strip().replace(" ", "")
		parsed = semester_to_academic_year_and_term(semester)
		if not parsed:
			msg = _("SEMESTER inválido")
			out["errors"].append({"row": row_num, "message": msg})
			_add_result(
				row=row_num,
				student_id=(row.get("ID") or "").strip(),
				course_input=(row.get("COURSE") or "").strip(),
				status="ErrorValidacion",
				detail=msg,
			)
			continue
		year, term_name = parsed
		term_label = SEMESTER_SUFFIX_TO_TERM.get(semester[-2:], "")
		course_code = (row.get("COURSE") or "").strip()
		course_frappe = _resolve_course(course_code)
		if not course_frappe:
			msg = _("Curso no existe: {0}").format(course_code)
			out["errors"].append({"row": row_num, "message": msg})
			_add_result(
				row=row_num,
				student_id=(row.get("ID") or "").strip(),
				course_input=course_code,
				academic_term=term_name,
				status="ErrorValidacion",
				detail=msg,
			)
			continue
		student_raw = (row.get("ID") or "").strip()
		student_name = get_student_name_by_id(student_raw)
		if not student_name:
			msg = _("Estudiante no encontrado: {0}").format(student_raw)
			out["errors"].append({"row": row_num, "message": msg})
			_add_result(
				row=row_num,
				student_id=student_raw,
				course_input=course_code,
				course=course_frappe,
				academic_term=term_name,
				status="ErrorValidacion",
				detail=msg,
			)
			continue
		enroll_date = coerce_enrollment_date_str(row.get("ENROLLMENT DATE")) or default_date
		key = (course_frappe, year, term_label)
		groups[key].append((row_num, student_name, student_raw, enroll_date))

	ce_meta = frappe.get_meta("Course Enrollment")
	has_custom_term = ce_meta.has_field("custom_academic_term")
	has_custom_year = ce_meta.has_field("custom_academic_year")

	created_or_updated_sg = set()
	total_ok = 0
	total_dup = 0

	group_keys = list(groups.keys())
	prog_total = max(len(data_rows), 1)
	prog_i = 0

	def _bump(msg):
		nonlocal prog_i
		prog_i += 1
		if progress_callback:
			progress_callback(min(prog_i, prog_total), prog_total, msg)

	for course_frappe, year, term_label in group_keys:
		term_name = f"{year} ({term_label})"
		rows = groups[(course_frappe, year, term_label)]

		student_names = [r[1] for r in rows]
		try:
			from edtools_core.course_enrollment_moodle import (
				enroll_moodle_instructors_from_student_group,
				prepare_moodle_course_for_enrollment_tool,
			)

			moodle_course_id = prepare_moodle_course_for_enrollment_tool(
				year,
				term_name,
				course_frappe,
				show_progress_msgs=False,
			)
			sg_name = ensure_student_group_for_course_import(
				course_frappe,
				year,
				term_name,
				term_label,
				year,
				student_names,
			)
			created_or_updated_sg.add(sg_name)
			enroll_moodle_instructors_from_student_group(
				sg_name,
				moodle_course_id,
				log_context="Course Enrollment Import",
			)
		except Exception as e:
			for row_num, __s, __raw, __d in rows:
				msg = _("Error Moodle/grupo: {0}").format(str(e)[:200])
				out["errors"].append(
					{
						"row": row_num,
						"message": msg,
					}
				)
				_add_result(
					row=row_num,
					student_id=__raw,
					student=__s,
					course_input=course_frappe,
					course=course_frappe,
					academic_term=term_name,
					status="ErrorMoodleGrupo",
					detail=msg,
				)
			continue

		seen_student_last_row: dict[str, tuple[int, str, str]] = {}
		for row_num, student_name, _student_raw, enroll_date in rows:
			seen_student_last_row[student_name] = (row_num, student_name, enroll_date)
		unique_rows = list(seen_student_last_row.values())

		for row_num, student_name, enroll_date in unique_rows:
			_bump(_("Inscribiendo: {0} — {1}").format(student_name, term_name))

			pe_name, pe_year, pe_err = get_unique_program_enrollment(student_name, year)
			if pe_err or not pe_name:
				msg = pe_err or _("Error PE")
				out["errors"].append({"row": row_num, "message": msg})
				_add_result(
					row=row_num,
					student_id=student_name,
					student=student_name,
					course_input=course_frappe,
					course=course_frappe,
					academic_term=term_name,
					status="ErrorPE",
					detail=msg,
				)
				continue

			filters: dict[str, Any] = {
				"student": student_name,
				"course": course_frappe,
				"docstatus": 1,
			}
			if has_custom_term and term_name:
				filters["custom_academic_term"] = term_name
			else:
				filters["program_enrollment"] = pe_name

			if frappe.db.exists("Course Enrollment", filters):
				total_dup += 1
				_add_result(
					row=row_num,
					student_id=student_name,
					student=student_name,
					course_input=course_frappe,
					course=course_frappe,
					academic_term=term_name,
					status="Duplicado",
					detail=_("Ya inscrito en este periodo."),
				)
				continue

			try:
				from edtools_core.moodle_sync import sync_student_enrollment_to_moodle

				sync_student_enrollment_to_moodle(
					student=student_name,
					academic_year=year,
					academic_term=term_name,
					course=course_frappe,
				)
			except Exception as moodle_err:
				msg = _("Moodle: {0}").format(str(moodle_err)[:180])
				out["errors"].append(
					{
						"row": row_num,
						"message": msg,
					}
				)
				_add_result(
					row=row_num,
					student_id=student_name,
					student=student_name,
					course_input=course_frappe,
					course=course_frappe,
					academic_term=term_name,
					status="ErrorMoodle",
					detail=msg,
				)
				frappe.log_error(
					title="Course Enrollment Import — Moodle",
					message=f"student={student_name} course={course_frappe}: {moodle_err}",
				)
				continue

			try:
				pe_doc = frappe.get_doc("Program Enrollment", pe_name)
				program = pe_doc.program
				ce_props: dict[str, Any] = {
					"doctype": "Course Enrollment",
					"student": student_name,
					"program": program,
					"course": course_frappe,
					"program_enrollment": pe_name,
					"enrollment_date": enroll_date,
				}
				if has_custom_year:
					ce_props["custom_academic_year"] = year
				if has_custom_term:
					ce_props["custom_academic_term"] = term_name
				enrollment = frappe.get_doc(ce_props)
				enrollment.insert(ignore_permissions=True)
				enrollment.submit()
				total_ok += 1
				detail = enrollment.name
				if pe_year and str(pe_year) != str(year):
					detail = _("{0} (PE {1})").format(enrollment.name, pe_year)
				_add_result(
					row=row_num,
					student_id=student_name,
					student=student_name,
					course_input=course_frappe,
					course=course_frappe,
					academic_term=term_name,
					status="Creado",
					detail=detail,
				)
			except Exception as e:
				msg = _plain_user_message_from_exception(e)
				out["errors"].append(
					{"row": row_num, "message": msg}
				)
				_add_result(
					row=row_num,
					student_id=student_name,
					student=student_name,
					course_input=course_frappe,
					course=course_frappe,
					academic_term=term_name,
					status="ErrorCE",
					detail=msg,
				)
				frappe.log_error(
					title="Course Enrollment Import — CE",
					message=f"student={student_name} course={course_frappe}: {e}",
				)

	out["summary"]["course_enrollments_created"] = total_ok
	out["summary"]["duplicates"] = total_dup
	out["summary"]["rows_with_errors"] = len(out["errors"])
	out["summary"]["rows_processed_ok"] = total_ok
	out["summary"]["student_groups_created_or_updated"] = len(created_or_updated_sg)
	out["results"] = [
		{
			**r,
			"student_id": html.escape(str(r.get("student_id", ""))),
			"student": html.escape(str(r.get("student", ""))),
			"course_input": html.escape(str(r.get("course_input", ""))),
			"course": html.escape(str(r.get("course", ""))),
			"academic_term": html.escape(str(r.get("academic_term", ""))),
			"status": html.escape(str(r.get("status", ""))),
			"detail": html.escape(str(r.get("detail", ""))),
		}
		for r in results
	]
	out["success"] = True
	return out
