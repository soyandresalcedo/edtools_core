# Copyright (c) 2026, EdTools and contributors
# Importación masiva de notas desde Excel/CSV: validación de formato, parser y procesamiento.

from __future__ import annotations

import csv
import io
import re
from typing import Any

import frappe
from frappe import _
from frappe.utils import flt

# Columnas requeridas en el archivo (coincidencia flexible por nombre)
REQUIRED_COLUMNS = ["ID", "SEMESTER", "COURSE", "FINAL GRADE"]
OPTIONAL_COLUMNS = ["FULL NAME", "COURSE TITLE"]

# Mapeo código SEMESTER (últimos 2 dígitos) -> nombre del periodo para Academic Term
# Formato del nombre en Frappe (Academic Term): "YYYY (Spring A)", "YYYY (Fall B)", etc.
SEMESTER_SUFFIX_TO_TERM = {
    "01": "Spring A",
    "02": "Spring B",
    "03": "Summer A",
    "04": "Summer B",
    "05": "Fall A",
    "06": "Fall B",
}


def _get_default_grading_scale() -> str | None:
    """
    Obtiene la escala por defecto de forma segura entre versiones:
    - Si Education Settings tiene default_grading_scale, la usa.
    - Si no existe el campo o falla la lectura, usa la primera Grading Scale disponible.
    """
    scale = None
    try:
        meta = frappe.get_meta("Education Settings")
        if meta and meta.has_field("default_grading_scale"):
            scale = frappe.db.get_single_value("Education Settings", "default_grading_scale")
    except Exception:
        scale = None

    if not scale:
        scales = frappe.get_all("Grading Scale", limit=1)
        scale = scales[0].name if scales else None

    return scale


def _resolve_file_path(file_path: str):
    """
    Convierte file_path (URL o ruta relativa) en ruta física del servidor.
    Soporta /files/ (público) y /private/files/ (privado) para evitar errores
    cuando el usuario sube el archivo como privado.
    Devuelve la ruta absoluta o None si no se puede resolver o el archivo no existe.
    """
    import os
    if not file_path or not isinstance(file_path, str):
        return None
    path = file_path.strip()
    if path.startswith("/private/files/") or path.startswith("private/files/"):
        rel = path.replace("/private/files/", "").replace("private/files/", "").lstrip("/")
        if not rel:
            return None
        resolved = frappe.get_site_path("private", "files", rel)
        return resolved if os.path.isfile(resolved) else None
    if path.startswith("/files/") or path.startswith("files/"):
        rel = path.replace("/files/", "").replace("files/", "").lstrip("/")
        resolved = frappe.get_site_path("public", "files", rel)
        return resolved if os.path.isfile(resolved) else None
    if os.path.isfile(path):
        return path
    if not os.path.isabs(path):
        resolved = frappe.get_site_path(path)
        return resolved if os.path.isfile(resolved) else None
    return None


def _normalize_header(h: str) -> str:
    """Normaliza nombre de columna para comparación (mayúsculas, espacios)."""
    if not h or not isinstance(h, str):
        return ""
    return " ".join(str(h).strip().upper().split())


def _find_column_index(headers: list[str], column_name: str) -> int | None:
    """Devuelve el índice de la columna que coincida con column_name (normalizado)."""
    target = _normalize_header(column_name)
    for i, h in enumerate(headers):
        if _normalize_header(h) == target:
            return i
    return None


def parse_file(file_path: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    Lee el archivo Excel o CSV y devuelve (headers_dict, rows).
    headers_dict: { "ID": 0, "SEMESTER": 1, ... } para acceso por nombre.
    rows: lista de dicts con keys = nombres de columna normalizados (ID, SEMESTER, etc.).
    Si el archivo no tiene filas de datos, rows puede estar vacío.
    """
    path_lower = (file_path or "").lower()
    if path_lower.endswith(".csv"):
        return _parse_csv(file_path)
    if path_lower.endswith(".xlsx") or path_lower.endswith(".xls"):
        return _parse_xlsx(file_path)
    return {}, []


def _parse_csv(file_path: str) -> tuple[dict[str, int], list[dict[str, Any]]]:
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows_list = list(reader)
    if not rows_list:
        return {}, []
    header_row = [str(c).strip() for c in rows_list[0]]
    col_index = {}
    for col_name in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
        idx = _find_column_index(header_row, col_name)
        if idx is not None:
            col_index[col_name] = idx
    data_rows = []
    for r in rows_list[1:]:
        row_dict = {}
        for name, idx in col_index.items():
            if idx < len(r):
                row_dict[name] = (r[idx] or "").strip()
            else:
                row_dict[name] = ""
        data_rows.append(row_dict)
    return col_index, data_rows


def _parse_xlsx(file_path: str) -> tuple[dict[str, int], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows_list = []
    for row in ws.iter_rows(values_only=True):
        rows_list.append([str(c) if c is not None else "" for c in row])
    wb.close()
    if not rows_list:
        return {}, []
    header_row = [str(c).strip() for c in rows_list[0]]
    col_index = {}
    for col_name in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
        idx = _find_column_index(header_row, col_name)
        if idx is not None:
            col_index[col_name] = idx
    data_rows = []
    for r in rows_list[1:]:
        row_dict = {}
        for name, idx in col_index.items():
            if idx < len(r):
                val = r[idx]
                row_dict[name] = (str(val).strip() if val is not None else "")
            else:
                row_dict[name] = ""
        data_rows.append(row_dict)
    return col_index, data_rows


def validate_format(
    file_path: str,
    grading_scale_name: str | None = None,
) -> tuple[bool, list[dict[str, Any]]]:
    """
    Verificación previa de formato. No crea ningún documento.
    Returns:
        (True, []) si todo está bien.
        (False, [{ "row": N, "message": "..." }, ...]) si hay errores de formato o columnas faltantes.
    """
    errors = []

    resolved = _resolve_file_path(file_path)
    if not resolved:
        errors.append({"row": None, "message": _("Se requiere un archivo Excel (.xlsx) o CSV.")})
        return False, errors
    file_path = resolved
    if not (file_path.lower().endswith(".csv") or file_path.lower().endswith(".xlsx") or file_path.lower().endswith(".xls")):
        errors.append({"row": None, "message": _("El archivo debe ser Excel (.xlsx) o CSV.")})
        return False, errors

    col_index, data_rows = parse_file(file_path)

    # 1) Columnas requeridas
    missing = []
    for col in REQUIRED_COLUMNS:
        if col not in col_index:
            missing.append(col)
    if missing:
        errors.append({
            "row": None,
            "message": _("Faltan columnas requeridas: {0}").format(", ".join(missing)),
        })
        return False, errors

    if not data_rows:
        errors.append({"row": None, "message": _("El archivo no tiene filas de datos.")})
        return False, errors

    # Criterio "Definitiva" requerido por el proceso de importación
    if not frappe.db.exists("Assessment Criteria", "Definitiva"):
        errors.append({
            "row": None,
            "message": _("No existe el criterio de evaluación 'Definitiva'. Créalo en Evaluación > Criterios de evaluación antes de importar."),
        })
        return False, errors

    # 2) Por cada fila: ID, SEMESTER, COURSE, FINAL GRADE
    semester_re = re.compile(r"^\d{6}$")
    for i, row in enumerate(data_rows):
        row_num = i + 2  # 1-based + header
        student_id = (row.get("ID") or "").strip()
        if not student_id:
            errors.append({"row": row_num, "message": _("Fila {0}: ID de estudiante no puede estar vacío.").format(row_num)})
        semester = (row.get("SEMESTER") or "").strip().replace(" ", "")
        if not semester:
            errors.append({"row": row_num, "message": _("Fila {0}: SEMESTER no puede estar vacío.").format(row_num)})
        elif not semester_re.match(semester):
            errors.append({"row": row_num, "message": _("Fila {0}: SEMESTER debe ser 6 dígitos (YYYY01 a YYYY06).").format(row_num)})
        else:
            suffix = semester[-2:]
            if suffix not in SEMESTER_SUFFIX_TO_TERM:
                errors.append({"row": row_num, "message": _("Fila {0}: SEMESTER debe terminar en 01-06 (ej. 202601).").format(row_num)})
        course = (row.get("COURSE") or "").strip()
        if not course:
            errors.append({"row": row_num, "message": _("Fila {0}: COURSE no puede estar vacío.").format(row_num)})
        grade = (row.get("FINAL GRADE") or "").strip()
        if not grade:
            errors.append({"row": row_num, "message": _("Fila {0}: FINAL GRADE no puede estar vacío.").format(row_num)})
        elif grading_scale_name and not _grade_value_valid(grade, grading_scale_name):
            errors.append({"row": row_num, "message": _("Fila {0}: FINAL GRADE '{1}' no es válido en la escala de calificaciones.").format(row_num, grade)})

    if errors:
        return False, errors
    return True, []


def _grade_value_valid(grade: str, grading_scale_name: str) -> bool:
    """Comprueba si grade es un número o una letra presente en la escala."""
    try:
        flt(grade)
        return True
    except (TypeError, ValueError):
        pass
    intervals = frappe.get_all(
        "Grading Scale Interval",
        filters={"parent": grading_scale_name},
        fields=["grade_code"],
    )
    for d in intervals or []:
        if (d.get("grade_code") or "").strip().upper() == grade.strip().upper():
            return True
    return False


def semester_to_academic_year_and_term(semester_code: str) -> tuple[str, str] | None:
    """
    Convierte código 202601 a (academic_year, academic_term_name).
    academic_year = "2026", academic_term_name = "2026 (Spring A)" (formato del DocType Academic Term).
    """
    semester_code = (semester_code or "").strip().replace(" ", "")
    if len(semester_code) != 6:
        return None
    suffix = semester_code[-2:]
    if suffix not in SEMESTER_SUFFIX_TO_TERM:
        return None
    year = semester_code[:4]
    term_label = SEMESTER_SUFFIX_TO_TERM[suffix]
    # Formato del nombre del Academic Term en Frappe: "2026 (Spring A)"
    term_name = f"{year} ({term_label})"
    return year, term_name


def letter_to_percentage(grading_scale_name: str, letter_grade: str) -> float | None:
    """
    Convierte una letra (A, A-, B+, etc.) al porcentaje mínimo de la escala.
    Usado para calcular score = (percentage/100) * max_score.
    """
    intervals = frappe.get_all(
        "Grading Scale Interval",
        filters={"parent": grading_scale_name},
        fields=["grade_code", "threshold"],
        order_by="threshold desc",
    )
    if not intervals:
        return None
    letter_grade = (letter_grade or "").strip().upper()
    for d in intervals or []:
        code = (d.get("grade_code") or "").strip().upper()
        if code == letter_grade:
            return flt(d.get("threshold"), 2)
    return None


def get_student_name_by_id(student_id: str) -> str | None:
    """
    Resuelve el estudiante por ID (columna ID del Excel).
    Busca por Student.name o por custom field que almacene el ID si existe.
    """
    if not student_id or not (student_id or "").strip():
        return None
    student_id = (student_id or "").strip()
    if frappe.db.exists("Student", student_id):
        return student_id
    # Opcional: custom field en Student para ID numérico (ej. student_id_number)
    meta = frappe.get_meta("Student")
    for f in meta.fields or []:
        if f.fieldtype in ("Data", "Int", "Small Text") and "id" in (f.fieldname or "").lower():
            try:
                name = frappe.db.get_value("Student", {f.fieldname: student_id}, "name")
                if name:
                    return name
            except Exception:
                pass
    return None


def _normalize_course_code(code: str) -> str:
    """Normaliza código de curso para búsqueda (quitar espacios extra, mantener formato)."""
    if not code:
        return ""
    return " ".join((code or "").strip().split())


def _expand_course_code_without_spaces(code: str) -> str | None:
    """
    Convierte código sin espacios a formato con espacio antes de los dígitos.
    Ej: STA530 -> STA 530, MIB650 -> MIB 650.
    Permite recibir STA530 cuando el Course tiene short_name = "STA 530".
    """
    if not code or not isinstance(code, str):
        return None
    code = code.strip().replace(" ", "")
    if not code:
        return None
    # Insertar espacio entre letras y dígitos: "STA530" -> "STA 530"
    m = re.match(r"^([A-Za-z]+)(\d.*)$", code)
    if m:
        return f"{m.group(1)} {m.group(2)}".strip()
    return None


def _resolve_course(course_code: str) -> str | None:
    """
    Devuelve el name del Course en Frappe si existe (por short_name o por name).
    Acepta códigos con o sin espacios: "STA 530" y "STA530" funcionan igual.
    """
    if not course_code:
        return None
    stripped = course_code.strip()
    normalized = _normalize_course_code(course_code)
    compact = (course_code or "").replace(" ", "")
    expanded = _expand_course_code_without_spaces(course_code)

    # 1) Buscar por short_name (si el DocType tiene el campo)
    meta = frappe.get_meta("Course")
    if meta.has_field("short_name"):
        values_to_try = [stripped, normalized, compact]
        if expanded and expanded not in values_to_try:
            values_to_try.append(expanded)
        for value in values_to_try:
            if not value:
                continue
            name = frappe.db.get_value("Course", {"short_name": value}, "name")
            if name:
                return name

    # 2) Búsqueda flexible: comparar short_name sin espacios (para DB con "STA 530", input "STA530")
    if meta.has_field("short_name") and compact:
        courses = frappe.get_all("Course", fields=["name", "short_name"])
        for c in courses or []:
            sn = (c.get("short_name") or "").strip().replace(" ", "")
            if sn == compact:
                return c.get("name")

    # 3) Buscar por name del registro
    for value in (normalized, stripped, compact, expanded):
        if value and frappe.db.exists("Course", value):
            return value
    return None


# Nombre del nodo raíz del árbol Assessment Group (puede existir en el sitio)
ASSESSMENT_GROUP_ROOT = "Todos los grupos de evaluación"


def get_or_create_assessment_group_leaf(academic_year: str, term_label: str) -> str | None:
    """
    Obtiene o crea el Assessment Group hoja "Nota definitiva - {term_label} - {year}".
    Crea si hace falta: raíz -> año -> periodo -> hoja.
    term_label ej. "Spring A", "Fall B".
    Devuelve el nombre del Assessment Group (la hoja) o None si falla.
    """
    year = (academic_year or "").strip()
    term_label = (term_label or "").strip()
    if not year or not term_label:
        return None
    leaf_name = f"Nota definitiva - {term_label} - {year}"
    if frappe.db.exists("Assessment Group", leaf_name):
        return leaf_name
    # Obtener o crear raíz
    root = frappe.db.get_value("Assessment Group", {"parent_assessment_group": ["is", "not set"]}, "name")
    if not root:
        root = frappe.db.get_value("Assessment Group", ASSESSMENT_GROUP_ROOT, "name")
    if not root:
        try:
            root_doc = frappe.new_doc("Assessment Group")
            root_doc.assessment_group_name = ASSESSMENT_GROUP_ROOT
            root_doc.parent_assessment_group = ASSESSMENT_GROUP_ROOT
            root_doc.is_group = 1
            root_doc.insert(ignore_permissions=True)
            frappe.db.commit()
            root = root_doc.name
        except Exception:
            root = ASSESSMENT_GROUP_ROOT
            if not frappe.db.exists("Assessment Group", root):
                return None
    # Nodo año
    year_node_name = year
    if not frappe.db.exists("Assessment Group", year_node_name):
        try:
            year_doc = frappe.new_doc("Assessment Group")
            year_doc.assessment_group_name = year_node_name
            year_doc.parent_assessment_group = root
            year_doc.is_group = 1
            year_doc.insert(ignore_permissions=True)
            frappe.db.commit()
        except Exception:
            pass
    # Nodo periodo del Assessment Group (nombre interno del árbol; no es el Academic Term)
    period_name = f"{year} - {term_label}"
    if not frappe.db.exists("Assessment Group", period_name):
        try:
            period_doc = frappe.new_doc("Assessment Group")
            period_doc.assessment_group_name = period_name
            period_doc.parent_assessment_group = year_node_name
            period_doc.is_group = 1
            period_doc.insert(ignore_permissions=True)
            frappe.db.commit()
        except Exception:
            pass
    # Hoja
    if not frappe.db.exists("Assessment Group", leaf_name):
        try:
            leaf_doc = frappe.new_doc("Assessment Group")
            leaf_doc.assessment_group_name = leaf_name
            leaf_doc.parent_assessment_group = period_name
            leaf_doc.is_group = 0
            leaf_doc.insert(ignore_permissions=True)
            frappe.db.commit()
        except Exception:
            return None
    return leaf_name


def get_or_create_student_group(
    course_name: str,
    academic_year: str,
    academic_term_name: str,
    student_names: list[str],
) -> str | None:
    """
    Obtiene o crea un Student Group para (course, academic_year, academic_term)
    con los estudiantes indicados. group_based_on = Course.
    Devuelve el nombre del Student Group o None.
    """
    if not course_name or not student_names:
        return None
    group_name = f"Grades - {course_name} - {academic_term_name}"
    if frappe.db.exists("Student Group", group_name):
        return group_name
    try:
        doc = frappe.new_doc("Student Group")
        doc.academic_year = academic_year
        doc.group_based_on = "Course"
        doc.student_group_name = group_name
        doc.academic_term = academic_term_name
        doc.course = course_name
        doc.program = None
        doc.max_strength = 0
        for i, stu in enumerate(student_names, 1):
            student_name_title = frappe.db.get_value("Student", stu, "student_name") or stu
            doc.append("students", {
                "student": stu,
                "student_name": student_name_title,
                "group_roll_number": i,
                "active": 1,
            })
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        return doc.name
    except Exception:
        frappe.db.rollback()
        return None


def get_or_create_assessment_plan(
    student_group_name: str,
    assessment_group_name: str,
    course_name: str,
    grading_scale_name: str,
    academic_term_name: str = "",
) -> str | None:
    """
    Obtiene o crea un Assessment Plan para el Student Group con criterio "Definitiva" 100.
    Devuelve el nombre del Assessment Plan o None.
    """
    existing = frappe.get_all(
        "Assessment Plan",
        filters={
            "student_group": student_group_name,
            "assessment_group": assessment_group_name,
            "docstatus": ["!=", 2],
        },
        limit=1,
    )
    if existing:
        return existing[0].name
    if not frappe.db.exists("Assessment Criteria", "Definitiva"):
        return None
    course_title = frappe.db.get_value("Course", course_name, "course_name") or course_name
    plan_title = f"Nota definitiva - {academic_term_name} - {course_title}" if academic_term_name else f"Nota definitiva - {course_title}"
    try:
        doc = frappe.new_doc("Assessment Plan")
        doc.assessment_name = plan_title
        doc.student_group = student_group_name
        doc.assessment_group = assessment_group_name
        doc.grading_scale = grading_scale_name
        doc.course = course_name
        doc.schedule_date = frappe.utils.getdate()
        doc.from_time = "09:00:00"
        doc.to_time = "10:00:00"
        doc.maximum_assessment_score = 100
        doc.append("assessment_criteria", {
            "assessment_criteria": "Definitiva",
            "maximum_score": 100,
        })
        doc.insert(ignore_permissions=True)
        doc.submit()
        frappe.db.commit()
        return doc.name
    except Exception:
        frappe.db.rollback()
        return None


def create_or_update_assessment_result(
    assessment_plan_name: str,
    student_name: str,
    score: float,
    grading_scale_name: str,
) -> tuple[str | None, str | None, bool, bool]:
    """
    Crea o actualiza el Assessment Result para (assessment_plan, student) con un solo criterio Definitiva y score.
    Devuelve (name, error_message, created, updated_submitted).
    updated_submitted=True cuando se actualizó un resultado que ya estaba presentado (sin cancelar).
    """
    try:
        from education.education.api import (
            get_assessment_result_doc,
            get_assessment_details,
            get_grade,
        )
    except ImportError:
        from education.education.education.api import (
            get_assessment_result_doc,
            get_assessment_details,
            get_grade,
        )

    details_list = get_assessment_details(assessment_plan_name)
    if not details_list or len(details_list) == 0:
        return None, _("El plan de evaluación no tiene criterios."), False, False
    criteria_name = details_list[0].get("assessment_criteria")
    if not criteria_name:
        return None, _("Criterio Definitiva no encontrado en el plan."), False, False

    plan = frappe.get_doc("Assessment Plan", assessment_plan_name)
    # Comprobar primero si ya existe resultado presentado: así no llamamos a get_assessment_result_doc
    # y evitamos el msgprint "Result already Submitted" de Education; mostramos el dato en el resumen.
    existing_submitted = frappe.get_all(
        "Assessment Result",
        filters={
            "student": student_name,
            "assessment_plan": assessment_plan_name,
            "docstatus": 1,
        },
        limit=1,
    )
    if existing_submitted:
        result_name = existing_submitted[0]["name"]
        grading_scale = plan.grading_scale or grading_scale_name
        score_val = flt(score, 2)
        percentage = (score_val / 100.0) * 100 if score_val else 0
        grade_letter = get_grade(grading_scale, percentage)
        details = frappe.get_all(
            "Assessment Result Detail",
            filters={"parent": result_name, "assessment_criteria": criteria_name},
            fields=["name"],
            limit=1,
        )
        if not details:
            return None, _("No se encontró el criterio Definitiva en el resultado existente."), False, False
        frappe.db.set_value("Assessment Result Detail", details[0]["name"], "score", score_val)
        frappe.db.set_value("Assessment Result Detail", details[0]["name"], "grade", grade_letter or "")
        frappe.db.set_value("Assessment Result", result_name, "total_score", score_val)
        frappe.db.set_value("Assessment Result", result_name, "grade", grade_letter or "")
        frappe.db.commit()
        frappe.clear_document_cache("Assessment Result", result_name)
        return result_name, None, False, True
    doc = get_assessment_result_doc(student_name, assessment_plan_name)
    if not doc:
        return None, _("No se pudo obtener o crear el documento de resultado."), False, False
    is_new = doc.get("__islocal", False)

    # Borrador: rellenar campos y guardar/presentar
    doc.assessment_plan = assessment_plan_name
    doc.student = student_name
    doc.student_group = plan.student_group
    doc.assessment_group = plan.assessment_group
    doc.grading_scale = plan.grading_scale or grading_scale_name
    doc.maximum_score = flt(plan.maximum_assessment_score, 2)

    doc.set("details", [])
    doc.append("details", {
        "assessment_criteria": criteria_name,
        "maximum_score": 100,
        "score": flt(score, 2),
    })
    doc.save(ignore_permissions=True)
    if doc.docstatus == 0:
        doc.submit()
    frappe.db.commit()
    return doc.name, None, is_new, False


# ---------------------------------------------------------------------------
# Importación individual (endpoint por estudiante)
# ---------------------------------------------------------------------------

# Campos esperados en el JSON del endpoint individual (equivalente al Excel)
SINGLE_GRADE_REQUIRED_FIELDS = ["student_id", "semester", "course", "final_grade"]
SINGLE_GRADE_OPTIONAL_FIELDS = ["full_name", "course_title", "grading_scale"]
SEMESTER_RE = re.compile(r"^\d{6}$")


def _normalize_single_grade_data(data: dict) -> dict:
    """Acepta nombres de Excel (ID, SEMESTER, etc.) o API (student_id, semester, etc.)."""
    aliases = {
        "ID": "student_id",
        "SEMESTER": "semester",
        "COURSE": "course",
        "FINAL GRADE": "final_grade",
        "FULL NAME": "full_name",
        "COURSE TITLE": "course_title",
        "GRADING SCALE": "grading_scale",
    }
    out = dict(data)
    for excel_name, api_name in aliases.items():
        if excel_name in out and api_name not in out:
            out[api_name] = out[excel_name]
    return out


def validate_grade_single_input(data: dict) -> tuple[bool, list[dict]]:
    """
    Valida los campos de una importación individual de nota.
    Devuelve (True, []) si todo es correcto, (False, errores) si hay problemas.

    Errores detallados por campo: faltante, formato incorrecto, valor inválido.
    Acepta nombres de Excel (ID, SEMESTER, COURSE, FINAL GRADE) o API (student_id, etc.).
    """
    errors: list[dict] = []
    if not isinstance(data, dict):
        return False, [{"field": None, "message": _("Se requiere un objeto JSON con los campos.")}]

    data = _normalize_single_grade_data(data)

    # 1) Campos requeridos faltantes
    for field in SINGLE_GRADE_REQUIRED_FIELDS:
        val = data.get(field)
        if val is None or (isinstance(val, str) and not str(val).strip()):
            errors.append({
                "field": field,
                "message": _("Campo obligatorio '{0}' no puede estar vacío.").format(field),
            })
    if errors:
        return False, errors

    student_id = str(data.get("student_id", "")).strip()
    semester = str(data.get("semester", "")).strip().replace(" ", "")
    course = str(data.get("course", "")).strip()
    final_grade = str(data.get("final_grade", "")).strip()
    grading_scale_name = (data.get("grading_scale") or "").strip() or None

    # 2) Formato SEMESTER: 6 dígitos, sufijo 01-06
    if not SEMESTER_RE.match(semester):
        errors.append({
            "field": "semester",
            "message": _("Debe ser 6 dígitos (YYYY01 a YYYY06, ej. 202601)."),
        })
    elif semester[-2:] not in SEMESTER_SUFFIX_TO_TERM:
        errors.append({
            "field": "semester",
            "message": _("Los dos últimos dígitos deben ser 01-06 (01=Spring A, 02=Spring B, etc.)."),
        })

    # 3) COURSE no vacío (el formato con/sin espacios se tolera en _resolve_course)
    if not course:
        errors.append({
            "field": "course",
            "message": _("No puede estar vacío. Ej: STA 530 o STA530."),
        })

    # 4) FINAL GRADE: no vacío, y si hay escala, válido
    if not final_grade:
        errors.append({
            "field": "final_grade",
            "message": _("No puede estar vacío. Debe ser un número (0-100) o una letra de la escala (A, B+, etc.)."),
        })
    else:
        scale = grading_scale_name or _get_default_grading_scale()
        if scale and not _grade_value_valid(final_grade, scale):
            errors.append({
                "field": "final_grade",
                "message": _("'{0}' no es válido. Debe ser un número o una letra de la escala de calificaciones.").format(final_grade),
            })

    # 5) Criterio Definitiva
    if not frappe.db.exists("Assessment Criteria", "Definitiva"):
        errors.append({
            "field": None,
            "message": _("No existe el criterio de evaluación 'Definitiva'. Créalo en Evaluación > Criterios de evaluación."),
        })

    # 6) Escala por defecto si no se indicó
    if not grading_scale_name:
        scale = _get_default_grading_scale()
        if not scale:
            errors.append({
                "field": "grading_scale",
                "message": _("No hay escala de calificaciones configurada. Indica 'grading_scale' o configura Education Settings."),
            })

    return len(errors) == 0, errors


def process_grade_single(data: dict) -> dict[str, Any]:
    """
    Procesa una sola nota (un estudiante, un curso, un período).
    Usa la misma lógica que process_grades pero para un único registro.

    Args:
        data: {
            "student_id": "EDU-STU-2025-02806" (obligatorio),
            "semester": "202601" (obligatorio, 6 dígitos),
            "course": "STA 530" o "STA530" (obligatorio),
            "final_grade": "85" o "A" (obligatorio),
            "full_name": "..." (opcional),
            "course_title": "..." (opcional),
            "grading_scale": "..." (opcional),
        }

    Returns:
        {
            "success": bool,
            "message": str,
            "validation_errors": [{"field": ..., "message": ...}],
            "assessment_result": nombre del documento creado/actualizado,
            "created": bool (True si nuevo, False si actualizado),
            "error_detail": str (si success=False por error de procesamiento),
        }
    """
    data = _normalize_single_grade_data(data)
    ok, validation_errors = validate_grade_single_input(data)
    if not ok:
        return {
            "success": False,
            "message": _("Errores de validación"),
            "validation_errors": validation_errors,
            "assessment_result": None,
            "created": False,
            "error_detail": "; ".join(e.get("message", "") for e in validation_errors),
        }

    student_id = str(data.get("student_id", "")).strip()
    semester = str(data.get("semester", "")).strip().replace(" ", "")
    course_code = str(data.get("course", "")).strip()
    final_grade_str = str(data.get("final_grade", "")).strip()
    grading_scale_name = (data.get("grading_scale") or "").strip() or None

    if not grading_scale_name:
        grading_scale_name = _get_default_grading_scale()

    parsed = semester_to_academic_year_and_term(semester)
    if not parsed:
        return {
            "success": False,
            "message": _("SEMESTER inválido"),
            "validation_errors": [],
            "assessment_result": None,
            "created": False,
            "error_detail": _("Semester debe ser 6 dígitos (ej. 202601)."),
        }

    # semester_to_academic_year_and_term ya devuelve (year, academic_term_name)
    # ejemplo: ("2026", "2026 (Spring A)")
    year, term_name = parsed
    term_label = SEMESTER_SUFFIX_TO_TERM.get(semester[-2:], "")

    course_frappe = _resolve_course(course_code)
    if not course_frappe:
        return {
            "success": False,
            "message": _("Curso no encontrado"),
            "validation_errors": [{"field": "course", "message": _("Curso '{0}' no existe. Prueba con o sin espacios (STA 530 / STA530).").format(course_code)}],
            "assessment_result": None,
            "created": False,
            "error_detail": _("Curso no existe: {0}").format(course_code),
        }

    student_name = get_student_name_by_id(student_id)
    if not student_name:
        return {
            "success": False,
            "message": _("Estudiante no encontrado"),
            "validation_errors": [{"field": "student_id", "message": _("Estudiante '{0}' no existe.").format(student_id)}],
            "assessment_result": None,
            "created": False,
            "error_detail": _("Estudiante no encontrado: {0}").format(student_id),
        }

    try:
        score = float(final_grade_str)
        score = flt(score, 2)
    except (TypeError, ValueError):
        pct = letter_to_percentage(grading_scale_name, final_grade_str)
        if pct is None:
            return {
                "success": False,
                "message": _("Calificación no válida"),
                "validation_errors": [{"field": "final_grade", "message": _("'{0}' no es un número ni una letra de la escala.").format(final_grade_str)}],
                "assessment_result": None,
                "created": False,
                "error_detail": _("Calificación no válida: {0}").format(final_grade_str),
            }
        score = (pct / 100.0) * 100

    leaf = get_or_create_assessment_group_leaf(year, term_label)
    if not leaf:
        return {
            "success": False,
            "message": _("Error al crear grupo de evaluación"),
            "validation_errors": [],
            "assessment_result": None,
            "created": False,
            "error_detail": _("No se pudo crear el grupo de evaluación para {0}.").format(term_name),
        }

    sg_name = get_or_create_student_group(course_frappe, year, term_name, [student_name])
    if not sg_name:
        return {
            "success": False,
            "message": _("Error al crear grupo de estudiantes"),
            "validation_errors": [],
            "assessment_result": None,
            "created": False,
            "error_detail": _("No se pudo crear el grupo de estudiantes."),
        }

    course_doc = frappe.get_cached_doc("Course", course_frappe)
    scale = getattr(course_doc, "default_grading_scale", None) or grading_scale_name
    ap_name = get_or_create_assessment_plan(sg_name, leaf, course_frappe, scale, academic_term_name=term_name)
    if not ap_name:
        return {
            "success": False,
            "message": _("Error al crear plan de evaluación"),
            "validation_errors": [],
            "assessment_result": None,
            "created": False,
            "error_detail": _("No se pudo crear el plan de evaluación."),
        }

    ar_name, err, created, _updated_submitted = create_or_update_assessment_result(
        ap_name, student_name, score, scale
    )
    if err:
        return {
            "success": False,
            "message": _("Error al guardar resultado"),
            "validation_errors": [],
            "assessment_result": None,
            "created": False,
            "error_detail": err,
        }

    return {
        "success": True,
        "message": _("Nota importada correctamente"),
        "validation_errors": [],
        "assessment_result": ar_name,
        "created": created,
        "error_detail": None,
    }


def process_grades(
    file_path: str,
    grading_scale_name: str | None = None,
    progress_callback: None | callable = None,
) -> dict[str, Any]:
    """
    Ejecuta la importación: validación previa y, si pasa, creación de grupos, planes y resultados.
    - Si la validación de formato falla, devuelve success=False y validation_errors (no crea nada).
    - Si la validación pasa, procesa filas; filas con error (estudiante no encontrado, etc.) se omiten
      y se anotan en errors; se crean resultados para el resto.
    Returns:
        {
            "success": bool,
            "validation_errors": [{"row": N, "message": "..."}],
            "summary": {"student_groups_created": 0, "assessment_plans_created": 0, "assessment_results_created": 0, "assessment_results_updated": 0, "assessment_results_updated_submitted": 0, "rows_processed": 0, "rows_with_errors": 0},
            "errors": [{"row": N, "message": "..."}],
        }
    """
    out = {
        "success": False,
        "validation_errors": [],
        "summary": {
            "student_groups_created": 0,
            "assessment_plans_created": 0,
            "assessment_results_created": 0,
            "assessment_results_updated": 0,
            "assessment_results_updated_submitted": 0,
            "rows_processed": 0,
            "rows_with_errors": 0,
        },
        "errors": [],
    }

    # 1) Validación previa
    ok, validation_errors = validate_format(file_path, grading_scale_name)
    if not ok:
        out["validation_errors"] = validation_errors
        return out

    resolved = _resolve_file_path(file_path)
    if not resolved:
        out["validation_errors"] = [{"row": None, "message": _("No se pudo leer el archivo.")}]
        return out

    col_index, data_rows = parse_file(resolved)
    if not data_rows:
        out["validation_errors"] = [{"row": None, "message": _("El archivo no tiene filas de datos.")}]
        return out

    # Grading scale por defecto: primer Course que usemos o primer escala en el sistema
    if not grading_scale_name:
        grading_scale_name = _get_default_grading_scale()
    if not grading_scale_name:
        out["validation_errors"] = [{"row": None, "message": _("No hay escala de calificaciones configurada.")}]
        return out

    # 2) Agrupar por (course, academic_year, academic_term)
    from collections import defaultdict
    groups = defaultdict(list)  # (course, year, term_label) -> [ (row_index, student_id, grade, course_code), ... ]
    for i, row in enumerate(data_rows):
        semester = (row.get("SEMESTER") or "").strip().replace(" ", "")
        parsed = semester_to_academic_year_and_term(semester)
        if not parsed:
            out["errors"].append({"row": i + 2, "message": _("SEMESTER inválido: {0}").format(row.get("SEMESTER"))})
            continue
        year, term_name = parsed
        term_label = SEMESTER_SUFFIX_TO_TERM.get(semester[-2:], "")
        course_code = (row.get("COURSE") or "").strip()
        course_frappe = _resolve_course(course_code)
        if not course_frappe:
            out["errors"].append({"row": i + 2, "message": _("Curso no existe: {0}").format(course_code)})
            continue
        student_id = (row.get("ID") or "").strip()
        student_name = get_student_name_by_id(student_id)
        if not student_name:
            out["errors"].append({"row": i + 2, "message": _("Estudiante no encontrado: {0}").format(student_id)})
            continue
        grade_str = (row.get("FINAL GRADE") or "").strip()
        if not grade_str:
            out["errors"].append({"row": i + 2, "message": _("FINAL GRADE no puede estar vacío.")})
            continue
        try:
            score = float(grade_str)
            score = flt(score, 2)
        except (TypeError, ValueError):
            pct = letter_to_percentage(grading_scale_name, grade_str)
            if pct is None:
                out["errors"].append({"row": i + 2, "message": _("Calificación no válida: {0}").format(grade_str)})
                continue
            score = (pct / 100.0) * 100
        key = (course_frappe, year, term_label)
        groups[key].append((i + 2, student_name, score, course_frappe))

    # 3) Por cada grupo: Assessment Group leaf, Student Group, Assessment Plan, Assessment Results
    created_sg = set()
    created_ap = set()
    total_processed = 0
    total_created = 0
    total_updated = 0
    total_updated_submitted = 0
    total_errors = len(out["errors"])

    for (course_frappe, year, term_label), rows in groups.items():
        term_name = f"{year} ({term_label})"
        if progress_callback:
            progress_callback(total_processed, len(data_rows), _("Procesando grupo {0} - {1}").format(course_frappe, term_name))
        leaf = get_or_create_assessment_group_leaf(year, term_label)
        if not leaf:
            for (row_num, __unused1, __unused2, __unused3) in rows:
                out["errors"].append({"row": row_num, "message": _("No se pudo crear el grupo de evaluación para {0}.").format(term_name)})
            continue
        student_names = list({r[1] for r in rows})
        sg_name = get_or_create_student_group(course_frappe, year, term_name, student_names)
        if not sg_name:
            for (row_num, __unused1, __unused2, __unused3) in rows:
                out["errors"].append({"row": row_num, "message": _("No se pudo crear el grupo de estudiantes.")})
            continue
        if sg_name not in created_sg:
            created_sg.add(sg_name)
        course_doc = frappe.get_cached_doc("Course", course_frappe)
        scale = getattr(course_doc, "default_grading_scale", None) or grading_scale_name
        ap_name = get_or_create_assessment_plan(sg_name, leaf, course_frappe, scale, academic_term_name=term_name)
        if not ap_name:
            for (row_num, __unused1, __unused2, __unused3) in rows:
                out["errors"].append({"row": row_num, "message": _("No se pudo crear el plan de evaluación.")})
            continue
        if ap_name not in created_ap:
            created_ap.add(ap_name)
        # Una sola fila por estudiante por grupo: la última en el archivo gana (evita que una fila con 0 o vacía sobrescriba la nota correcta).
        seen_student = {}
        for row_num, student_name, score, __unused_course in rows:
            seen_student[student_name] = (row_num, student_name, score, __unused_course)
        unique_rows = list(seen_student.values())
        for row_num, student_name, score, __unused_course in unique_rows:
            if progress_callback:
                progress_callback(total_processed, len(data_rows), _("Procesando resultado: {0}").format(student_name))
            ar_name, err, created, updated_submitted = create_or_update_assessment_result(ap_name, student_name, score, scale)
            if err:
                out["errors"].append({"row": row_num, "message": err})
            else:
                total_processed += 1
                if created:
                    total_created += 1
                else:
                    total_updated += 1
                    if updated_submitted:
                        total_updated_submitted += 1

    out["summary"]["student_groups_created"] = len(created_sg)
    out["summary"]["assessment_plans_created"] = len(created_ap)
    out["summary"]["assessment_results_created"] = total_created
    out["summary"]["assessment_results_updated"] = total_updated
    out["summary"]["assessment_results_updated_submitted"] = total_updated_submitted
    out["summary"]["rows_processed"] = total_processed
    out["summary"]["rows_with_errors"] = len(out["errors"])
    out["success"] = True
    return out
